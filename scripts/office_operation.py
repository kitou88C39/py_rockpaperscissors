#excel
#sample1
import openpyxl as px

wb = px.Workbook()
wb.save('excel_sample.xlsx')

#sample2
import openpyxl as px

wb = px.load.Workbook('excel_sample.xlsx')
ws = wb.active
ws['A1'] = 'Hello'
ws.title = 'sample_title1'
wb.save('excel_sample.xlsx')

#sample3
import openpyxl as px

wb = px.load.Workbook('excel_sample.xlsx')
ws = wb.create_sheet()
ws.title = 'sample_title1'

ws.append(['Data1'])

for i in range(10):
    ws.append([i + 2])
value = px.chart.Reference(ws, min_col=1, min_row=1, max_col=1, maz_row=10)
chart = px.chart.BarChart()
chart.add_data(value, title_from_data=True)
ws.add_chart(chart, "E15")
wb.save('excel_sample.xlsx')

#word
#sample1

import docx

document = docx.Document()
document.save('word_sample.docx')